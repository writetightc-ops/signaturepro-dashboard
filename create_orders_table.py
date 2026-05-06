"""
Создаёт новый облегчённый Excel-файл заказов SignaturePro_Заказы_NEW.xlsx.
Импортирует данные из старого файла (Апрель_2026).
Структура: 19 колонок вместо 30+ — только ввод данных, без зарплатных расчётов.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OLD_FILE = os.path.join(BASE_DIR, "SignaturePro _ заказы 2026 (1).xlsx")
NEW_FILE = os.path.join(BASE_DIR, "SignaturePro_Заказы_NEW.xlsx")

# ─── Колонки нового файла ────────────────────────────────────────────────────
NEW_HEADERS = [
    "Дата поступления",   # A
    "Статус",             # B  (True = завершён)
    "Менеджер",           # C
    "Каллиграф",          # D
    "ФИО клиента",        # E
    "Тариф",              # F
    "Варианты",           # G  — дата готовности вариантов
    "ОС 1",               # H  — обратная связь 1
    "Правка 1",           # I
    "ОС 2",               # J
    "Правка 2",           # K
    "ОС 3",               # L
    "Правка 3",           # M
    "ОС 4",               # N
    "Правка 4",           # O
    "Обучение",           # P  — дата готовности обучения
    "Ссылка",             # Q  — ссылка на обучение
    "Дата завершения",    # R  — дата отправки обучения клиенту
    "Заметки",            # S
]

# Маппинг из старых колонок (индексы в Апрель_2026, 0-based, строка данных)
OLD_IDX = {
    "поступление": 2,
    "статус":       1,
    "менеджер":     0,
    "каллиграф":    3,
    "клиент":       4,
    "тариф":        5,
    "варианты":     6,
    "ос1":          7,
    "правка1":      8,
    "ос2":          9,
    "правка2":     10,
    "ос3":         11,
    "правка3":     12,
    "ос4":         13,
    "правка4":     14,
    "обучение":    21,
    "ссылка":      23,
    "завершение":  24,
    "заметки":     25,
}

# ─── Цветовая схема ──────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1E3A5F"   # тёмно-синий — шапка
CLR_HEADER_FG   = "FFFFFF"   # белый текст
CLR_DONE        = "D6EAD4"   # зелёный — завершённый заказ
CLR_IN_PROGRESS = "FFF9E6"   # жёлтый — в работе
CLR_ALT_ROW     = "F5F8FF"   # голубоватый — чередование строк
CLR_DATE_COL    = "EBF0FF"   # светло-синий — колонки с датами этапов
CLR_BORDER      = "B0BEC5"

DATE_COLS = [0, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 17]  # 0-based индексы дат


def make_border(color=CLR_BORDER):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_header_cell(cell, text):
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, color=CLR_HEADER_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border("FFFFFF")


def build_sheet(ws, sheet_name, rows_data):
    """Создаёт и стилизует лист заказов."""
    ws.title = sheet_name

    # ── Заголовок ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    for col_idx, header in enumerate(NEW_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        style_header_cell(cell, header)

    # ── Ширины колонок ───────────────────────────────────────────────────────
    col_widths = {
        1: 14,   # Дата поступления
        2: 9,    # Статус
        3: 18,   # Менеджер
        4: 16,   # Каллиграф
        5: 24,   # ФИО клиента
        6: 10,   # Тариф
        7: 12,   # Варианты
        8: 12,   # ОС 1
        9: 12,   # Правка 1
        10: 12,  # ОС 2
        11: 12,  # Правка 2
        12: 12,  # ОС 3
        13: 12,  # Правка 3
        14: 12,  # ОС 4
        15: 12,  # Правка 4
        16: 12,  # Обучение
        17: 30,  # Ссылка
        18: 14,  # Дата завершения
        19: 30,  # Заметки
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Данные ───────────────────────────────────────────────────────────────
    for row_num, row in enumerate(rows_data, start=2):
        is_done = row[1] is True
        bg = CLR_DONE if is_done else CLR_IN_PROGRESS if row[6] else CLR_ALT_ROW

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = make_border()
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                horizontal="center" if col_idx not in (3, 4, 5, 19) else "left",
                vertical="center",
            )
            # Дата-формат
            if col_idx - 1 in DATE_COLS and isinstance(value, datetime):
                cell.number_format = "DD.MM.YY"
            # Фон: ссылка и заметки — белые; остальное — по статусу
            if col_idx in (17, 19):
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
                if col_idx == 17:
                    cell.font = Font(name="Calibri", size=9, color="1155CC", underline="single")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.fill = PatternFill("solid", fgColor=bg)

        ws.row_dimensions[row_num].height = 16

    # ── Заморозка шапки ──────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Автофильтр ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(NEW_HEADERS))}1"

    # ── Валидация: статус ─────────────────────────────────────────────────────
    dv_status = DataValidation(
        type="list", formula1='"TRUE,FALSE"', allow_blank=True,
        showDropDown=False, error="Введите TRUE или FALSE"
    )
    ws.add_data_validation(dv_status)
    dv_status.add(f"B2:B{max(len(rows_data) + 50, 200)}")

    print(f"  Лист «{sheet_name}»: {len(rows_data)} строк")


def build_handbook_sheet(wb):
    """Справочник: тарифы, сотрудники, ставки."""
    ws = wb.create_sheet("Справочник")
    ws.sheet_properties.tabColor = "2196F3"

    def h(cell, text):
        cell.value = text
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Тарифы и ставки каллиграфа ────────────────────────────────────────────
    ws.cell(1, 1).value = "ТАРИФЫ И МОТИВАЦИЯ КАЛЛИГРАФОВ"
    ws.cell(1, 1).font = Font(bold=True, size=11, color="1E3A5F")
    ws.merge_cells("A1:G1")

    tariff_headers = ["Тариф", "Варианты", "Обучение", "Бонус без правок",
                       "Тарифный бонус", "Итог без правок", "Стоимость 1 правки"]
    for ci, hdr in enumerate(tariff_headers, 1):
        h(ws.cell(2, ci), hdr)

    tariff_data = [
        ("E",      150, 200, 100, 0,   450,  75),
        ("ST",     150, 200, 100, 0,   450,  75),
        ("E_FAST", 150, 200, 150, 300, 800,  75),
        ("OPTNEW", 500, 200, 200, 0,   900,  75),
        ("PRNEW",  500, 200, 250, 300, 1250, 75),
    ]
    for ri, row in enumerate(tariff_data, 3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor="F0F4FF" if ri % 2 == 0 else "FFFFFF")

    # ── Бонус менеджера за завершённый заказ ─────────────────────────────────
    ws.cell(9, 1).value = "БОНУС МЕНЕДЖЕРА (за завершённый заказ)"
    ws.cell(9, 1).font = Font(bold=True, size=11, color="1E3A5F")
    ws.merge_cells("A9:G9")

    mgr_headers = ["Тариф", "Бонус менеджера, ₽"]
    for ci, hdr in enumerate(mgr_headers, 1):
        h(ws.cell(10, ci), hdr)

    mgr_data = [
        ("E",      200),
        ("ST",     200),
        ("E_FAST", 200),
        ("OPTNEW", 300),
        ("PRNEW",  300),
    ]
    for ri, row in enumerate(mgr_data, 11):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(horizontal="center")

    # ── Сотрудники ────────────────────────────────────────────────────────────
    ws.cell(17, 1).value = "СОТРУДНИКИ"
    ws.cell(17, 1).font = Font(bold=True, size=11, color="1E3A5F")
    ws.merge_cells("A17:E17")

    emp_headers = ["Имя", "Роль", "Грейд", "Коэффициент", "Активен"]
    for ci, hdr in enumerate(emp_headers, 1):
        h(ws.cell(18, ci), hdr)

    emp_data = [
        ("Марьям",          "Каллиграф", "Старший", 1.0, True),
        ("Лена Вовина",     "Каллиграф", "Старший", 1.0, True),
        ("Катерина Попова", "Каллиграф", "Старший", 1.0, True),
        ("Катя Дорожкина",  "Каллиграф", "Старший", 1.0, True),
        ("Ольга Струкова",  "Менеджер",  "-",        1.0, True),
        ("Мария Тимофеева", "Менеджер",  "-",        1.0, True),
    ]
    for ri, row in enumerate(emp_data, 19):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(horizontal="center")

    for col, width in zip("ABCDEFG", [14, 12, 10, 14, 8, 18, 18]):
        ws.column_dimensions[col].width = width

    print("  Лист «Справочник» создан")


def load_old_data():
    """Читает Апрель_2026 из старого файла, возвращает список строк нового формата."""
    print(f"Читаю старый файл: {OLD_FILE}")
    wb = openpyxl.load_workbook(OLD_FILE, data_only=True)
    ws = wb["Апрель_2026"]

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Пропуск пустых строк
        if all(v is None for v in row):
            continue
        # Пропуск разделителей-шаблонов
        if row[0] == "-" or row[2] is None:
            continue

        def get(idx):
            return row[idx] if idx < len(row) else None

        new_row = [
            get(OLD_IDX["поступление"]),   # A Дата поступления
            get(OLD_IDX["статус"]),         # B Статус
            get(OLD_IDX["менеджер"]),       # C Менеджер
            get(OLD_IDX["каллиграф"]),      # D Каллиграф
            get(OLD_IDX["клиент"]),         # E ФИО клиента
            get(OLD_IDX["тариф"]),          # F Тариф
            get(OLD_IDX["варианты"]),       # G Варианты
            get(OLD_IDX["ос1"]),            # H ОС 1
            get(OLD_IDX["правка1"]),        # I Правка 1
            get(OLD_IDX["ос2"]),            # J ОС 2
            get(OLD_IDX["правка2"]),        # K Правка 2
            get(OLD_IDX["ос3"]),            # L ОС 3
            get(OLD_IDX["правка3"]),        # M Правка 3
            get(OLD_IDX["ос4"]),            # N ОС 4
            get(OLD_IDX["правка4"]),        # O Правка 4
            get(OLD_IDX["обучение"]),       # P Обучение
            get(OLD_IDX["ссылка"]),         # Q Ссылка
            get(OLD_IDX["завершение"]),     # R Дата завершения
            get(OLD_IDX["заметки"]),        # S Заметки
        ]
        rows.append(new_row)

    print(f"  Загружено {len(rows)} заказов из Апрель_2026")
    return rows


def build_template_sheet(wb):
    """Пустой шаблон для нового месяца."""
    ws = wb.create_sheet("Шаблон")
    ws.sheet_properties.tabColor = "4CAF50"
    ws.row_dimensions[1].height = 36
    for col_idx, header in enumerate(NEW_HEADERS, start=1):
        style_header_cell(ws.cell(row=1, column=col_idx), header)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(NEW_HEADERS))}1"

    col_widths = [14,9,18,16,24,10,12,12,12,12,12,12,12,12,12,12,30,14,30]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.cell(2, 1).value = "← Скопируй этот лист и переименуй: Май_2026, Июнь_2026 и т.д."
    ws.cell(2, 1).font = Font(italic=True, color="888888", size=9)
    ws.merge_cells("A2:S2")
    print("  Лист «Шаблон» создан")


def main():
    print("\n=== Создание новой таблицы заказов ===\n")

    # Загружаем данные из старого файла
    april_rows = load_old_data()

    # Создаём новый файл
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # удаляем пустой лист по умолчанию

    # Лист с данными апреля
    ws_april = wb.create_sheet("Апрель_2026")
    ws_april.sheet_properties.tabColor = "FF6B35"
    build_sheet(ws_april, "Апрель_2026", april_rows)

    # Шаблон и справочник
    build_template_sheet(wb)
    build_handbook_sheet(wb)

    wb.save(NEW_FILE)
    print(f"\nФайл сохранён: {NEW_FILE}")
    print(f"   Колонок в новой таблице: {len(NEW_HEADERS)} (было 30–54)")
    print(f"   Заказов импортировано:   {len(april_rows)}")
    print("\nЧто убрано из старой таблицы:")
    print("  • ОС 5–7, Правка 5–7 (крайне редко нужны — добавишь вручную при необходимости)")
    print("  • Анализ (не нужен для ввода данных)")
    print("  • Доп. поля 1–5")
    print("  • Все зарплатные расчёты (теперь в дашборде)")
    print("\nЗапусти dashboard.py для просмотра ЗП в браузере.")


if __name__ == "__main__":
    main()
