"""
SignaturePro — Дашборд ЗП
Запуск: python dashboard.py
Открыть: http://localhost:5000

Режимы данных:
  1. Локальный Excel (по умолчанию): читает SignaturePro_Заказы_NEW.xlsx
  2. Google Sheets: заполни GOOGLE_SHEETS_URL в config.py
"""

from flask import Flask, render_template, request, jsonify
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os

app = Flask(__name__, template_folder="templates")

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "SignaturePro_Заказы_NEW.xlsx")

# ─── Подключение к источнику данных (config.py или переменные окружения) ──────
#
# Если приложение запущено на сервере (Render и т.п.),
# credentials можно передать через переменную окружения GOOGLE_CREDENTIALS_B64
# (base64-строка из credentials.json).
#
def _prepare_cloud_credentials():
    """Создаёт credentials.json из переменной окружения (для хостинга на Render)."""
    b64 = os.environ.get("GOOGLE_CREDENTIALS_B64", "")
    if not b64:
        return None
    import base64, json, tempfile
    try:
        data = base64.b64decode(b64).decode("utf-8")
        json.loads(data)  # проверяем что это валидный JSON
        tmp = os.path.join(BASE_DIR, "_credentials_cloud.json")
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(data)
        return tmp
    except Exception as e:
        print(f"  Ошибка декодирования GOOGLE_CREDENTIALS_B64: {e}")
        return None

try:
    from config import GOOGLE_SHEETS_URL, CREDENTIALS_FILE, CACHE_SECONDS
    _CREDS_PATH = os.path.join(BASE_DIR, CREDENTIALS_FILE)
except ImportError:
    GOOGLE_SHEETS_URL = os.environ.get("GOOGLE_SHEETS_URL", "")
    CREDENTIALS_FILE  = "credentials.json"
    CACHE_SECONDS     = 45
    _CREDS_PATH       = os.path.join(BASE_DIR, CREDENTIALS_FILE)

# Если credentials нет локально — пробуем из переменной окружения
if not os.path.exists(_CREDS_PATH):
    cloud_path = _prepare_cloud_credentials()
    if cloud_path:
        _CREDS_PATH = cloud_path

USE_GSHEETS = bool(GOOGLE_SHEETS_URL and os.path.exists(_CREDS_PATH))

if USE_GSHEETS:
    try:
        import gsheets as _gs
        print(f"  Источник данных: Google Sheets")
        print(f"  URL: {GOOGLE_SHEETS_URL[:60]}...")
    except ImportError as e:
        print(f"  Ошибка импорта gsheets: {e}")
        USE_GSHEETS = False
else:
    print("  Источник данных: локальный Excel")

# ─── Мотивация каллиграфов (таблица «Мотивация каллиграфы новая») ────────────
# варианты       : выплата когда проставлена дата «Варианты»
# обучение       : выплата когда проставлена дата «Обучение»
# тарифный_бонус : выплата при завершении заказа (всегда, если > 0)
# бонус_без_правок: выплата при завершении ТОЛЬКО если ни одной правки по заказу
# правка         : выплата за каждую правку (когда проставлена дата «Правка N»)
CALLIGRAPHER_RATES = {
    "E":      {"варианты": 150, "обучение": 200, "бонус_без_правок": 100, "тарифный_бонус":   0, "правка": 75},
    "ST":     {"варианты": 150, "обучение": 200, "бонус_без_правок": 100, "тарифный_бонус":   0, "правка": 75},
    "E_FAST": {"варианты": 150, "обучение": 200, "бонус_без_правок": 150, "тарифный_бонус": 300, "правка": 75},
    "OPTNEW": {"варианты": 500, "обучение": 200, "бонус_без_правок": 200, "тарифный_бонус":   0, "правка": 75},
    "PRNEW":  {"варианты": 500, "обучение": 200, "бонус_без_правок": 250, "тарифный_бонус": 300, "правка": 75},
}

# ─── Бонус менеджера за каждый завершённый заказ ────────────────────────────
MANAGER_RATES = {
    "E":      200,
    "ST":     200,
    "E_FAST": 200,
    "OPTNEW": 300,
    "PRNEW":  300,
}

DEFAULT_GRADES = {
    "Марьям":          1.0,
    "Лена Вовина":     1.0,
    "Катерина Попова": 1.0,
    "Катя Дорожкина":  1.0,
    "Ольга Струкова":  1.0,
    "Мария Тимофеева": 1.0,
}

SKIP_SHEETS = {"Справочник", "Шаблон", "import", "export", "Дисциплина",
               "ЗП отчет месяц", "Бонусы", "Рейты", "Сотрудники",
               "Статистика", "Импорт", "Sheet"}

NEW_HEADERS = [
    "Дата поступления", "Статус", "Менеджер", "Каллиграф", "ФИО клиента",
    "Тариф", "Варианты", "ОС 1", "Правка 1", "ОС 2", "Правка 2",
    "ОС 3", "Правка 3", "ОС 4", "Правка 4", "Обучение",
    "Ссылка", "Дата завершения", "Заметки",
]

# ─── Утилиты ─────────────────────────────────────────────────────────────────

def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def fmt_date(d):
    if d is None:
        return ""
    d = to_date(d)
    return d.strftime("%d.%m.%y") if d else ""


def detect_order_sheets(wb):
    return [s for s in wb.sheetnames if s not in SKIP_SHEETS]


def load_employees(wb):
    grades = dict(DEFAULT_GRADES)
    if "Справочник" not in wb.sheetnames:
        return grades
    for row in wb["Справочник"].iter_rows(min_row=19, values_only=True):
        name, _, _, coeff, *_ = (list(row) + [None] * 5)[:5]
        if name and isinstance(name, str) and isinstance(coeff, (int, float)):
            grades[name.strip()] = float(coeff)
    return grades


def load_orders_from_sheet(ws):
    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or (row[0] is None and (len(row) < 3 or row[2] is None)):
            continue
        if isinstance(row[0], str) and row[0].startswith("←"):
            continue
        r = list(row) + [None] * 20
        orders.append({
            "поступление": to_date(r[0]),
            "статус":      r[1],
            "менеджер":    str(r[2]).strip() if r[2] else "",
            "каллиграф":   str(r[3]).strip() if r[3] else "",
            "клиент":      str(r[4]).strip() if r[4] else "",
            "тариф":       str(r[5]).strip().upper() if r[5] else "",
            "варианты":    to_date(r[6]),
            "ос1":         to_date(r[7]),
            "правка1":     to_date(r[8]),
            "ос2":         to_date(r[9]),
            "правка2":     to_date(r[10]),
            "ос3":         to_date(r[11]),
            "правка3":     to_date(r[12]),
            "ос4":         to_date(r[13]),
            "правка4":     to_date(r[14]),
            "обучение":    to_date(r[15]),
            "ссылка":      r[16] if r[16] else "",
            "завершение":  to_date(r[17]),
            "заметки":     str(r[18]).strip() if r[18] else "",
        })
    return orders


def calc_order_earnings(order, date_from, date_to):
    """
    Считает заработок каллиграфа и менеджера по одному заказу
    за период [date_from, date_to].

    Логика:
    - Каждый этап (варианты, правки, обучение) засчитывается, если его дата
      попадает в период — независимо от того, завершён ли заказ.
    - Тарифный бонус и бонус_без_правок — только при завершении заказа в периоде.
    - Бонус_без_правок — только если на заказе ВООБЩЕ нет правок (не только в периоде).
    - Менеджерский бонус — при завершении заказа в периоде.
    """
    tariff = order["тариф"]
    rates = CALLIGRAPHER_RATES.get(tariff)
    if not rates:
        return {}, 0.0, 0.0

    def in_p(d):
        return d is not None and date_from <= d <= date_to

    cal = {}

    # Варианты
    if in_p(order["варианты"]):
        cal["варианты"] = rates["варианты"]

    # Правки (за каждую, попавшую в период)
    правки_sum = 0
    for k in ("правка1", "правка2", "правка3", "правка4"):
        if in_p(order[k]):
            правки_sum += rates["правка"]
    if правки_sum:
        cal["правки"] = правки_sum

    # Обучение
    if in_p(order["обучение"]):
        cal["обучение"] = rates["обучение"]

    # Завершение: бонусы — только если завершение попало в период
    mgr_bonus = 0.0
    if in_p(order["завершение"]):
        if rates["тарифный_бонус"] > 0:
            cal["тарифный_бонус"] = rates["тарифный_бонус"]

        # Бонус без правок — смотрим на ВСЕ правки по заказу, не только в периоде
        total_правок_по_заказу = sum(
            1 for k in ("правка1", "правка2", "правка3", "правка4")
            if order[k] is not None
        )
        if total_правок_по_заказу == 0:
            cal["бонус_без_правок"] = rates["бонус_без_правок"]

        mgr_bonus = float(MANAGER_RATES.get(tariff, 0))

    cal_total = sum(cal.values())
    return cal, cal_total, mgr_bonus


def _load_all_orders_excel():
    """Читает все заказы из локального Excel. Возвращает (orders, sheets, grades)."""
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(
            f"Файл не найден: {EXCEL_FILE}\n"
            "Запусти: python create_orders_table.py"
        )
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    grades = load_employees(wb)
    order_sheets = detect_order_sheets(wb)
    all_orders = []
    for sh in order_sheets:
        for o in load_orders_from_sheet(wb[sh]):
            o["_sheet"] = sh
            all_orders.append(o)
    return all_orders, order_sheets, grades


def _load_all_orders_gsheets():
    """Читает все заказы из Google Sheets. Возвращает (orders, sheets, grades)."""
    orders, sheets = _gs.read_all_orders(GOOGLE_SHEETS_URL, _CREDS_PATH, CACHE_SECONDS)
    return orders, sheets, dict(DEFAULT_GRADES)


def calculate_salary(date_from, date_to):
    """
    Считает ЗП всех сотрудников за период date_from..date_to.
    Читает ВСЕ листы заказов — незавершённые заказы находятся на любом листе.
    Автоматически выбирает источник данных (Excel или Google Sheets).
    """
    try:
        if USE_GSHEETS:
            all_orders, order_sheets, grades = _load_all_orders_gsheets()
        else:
            all_orders, order_sheets, grades = _load_all_orders_excel()
    except (FileNotFoundError, Exception) as e:
        return {"error": str(e)}

    results = {}

    def ensure(name, role):  # noqa: E306
        if name not in results:
            results[name] = {
                "name": name, "role": role,
                "coefficient": grades.get(name, 1.0),
                "total": 0.0,
                "breakdown": {},
                "orders_count": 0,
                "orders_done": 0,
                "orders_detail": [],
            }

    for order in all_orders:
        cal_name = order["каллиграф"]
        mgr_name = order["менеджер"]
        if not cal_name and not mgr_name:
            continue

        cal_bd, cal_total, mgr_bonus = calc_order_earnings(order, date_from, date_to)

        # ── Каллиграф ───────────────────────────────────────────────────────
        if cal_name and cal_total > 0:
            ensure(cal_name, "Каллиграф")
            coeff = grades.get(cal_name, 1.0)
            weighted = cal_total * coeff
            results[cal_name]["total"] += weighted
            results[cal_name]["orders_count"] += 1
            if order["завершение"] and date_from <= order["завершение"] <= date_to:
                results[cal_name]["orders_done"] += 1
            for k, v in cal_bd.items():
                bd = results[cal_name]["breakdown"]
                bd[k] = bd.get(k, 0) + v * coeff
            results[cal_name]["orders_detail"].append({
                "клиент":      order["клиент"],
                "тариф":       order["тариф"],
                "поступление": fmt_date(order["поступление"]),
                "завершение":  fmt_date(order["завершение"]),
                "варианты":    fmt_date(order["варианты"]),
                "правок":      sum(1 for k in ("правка1","правка2","правка3","правка4") if order[k]),
                "обучение":    fmt_date(order["обучение"]),
                "заработок":   round(cal_total * coeff, 2),
                "breakdown":   {k: round(v * coeff, 2) for k, v in cal_bd.items()},
                "лист":        order["_sheet"],
            })

        # ── Менеджер ─────────────────────────────────────────────────────────
        if mgr_name and mgr_bonus > 0:
            ensure(mgr_name, "Менеджер")
            results[mgr_name]["total"] += mgr_bonus
            bd = results[mgr_name]["breakdown"]
            bd["бонус_за_заказ"] = bd.get("бонус_за_заказ", 0) + mgr_bonus
            results[mgr_name]["orders_count"] += 1
            results[mgr_name]["orders_done"] += 1
            results[mgr_name]["orders_detail"].append({
                "клиент":      order["клиент"],
                "тариф":       order["тариф"],
                "поступление": fmt_date(order["поступление"]),
                "завершение":  fmt_date(order["завершение"]),
                "варианты":    "",
                "правок":      sum(1 for k in ("правка1","правка2","правка3","правка4") if order[k]),
                "обучение":    "",
                "заработок":   round(mgr_bonus, 2),
                "breakdown":   {"бонус_за_заказ": round(mgr_bonus, 2)},
                "лист":        order["_sheet"],
            })

    for emp in results.values():
        emp["total"] = round(emp["total"], 2)
        emp["breakdown"] = {k: round(v, 2) for k, v in emp["breakdown"].items() if v > 0}

    return {
        "employees": results,
        "period_from": date_from.isoformat(),
        "period_to":   date_to.isoformat(),
        "sheets_used": order_sheets,
        "total_orders": len(all_orders),
    }


# ─── Создание нового листа-месяца ────────────────────────────────────────────

def _border():
    s = Side(style="thin", color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)


def create_month_sheet(wb, sheet_name):
    """Добавляет новый лист с заголовками по шаблону."""
    ws = wb.create_sheet(sheet_name)
    ws.row_dimensions[1].height = 36
    col_widths = [14,9,18,16,24,10,12,12,12,12,12,12,12,12,12,12,30,14,30]
    for ci, (hdr, w) in enumerate(zip(NEW_HEADERS, col_widths), 1):
        cell = ws.cell(1, ci, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(NEW_HEADERS))}1"
    return ws


# ─── Flask маршруты ──────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/meta")
def api_meta():
    """Возвращает список листов и источник данных."""
    source = "gsheets" if USE_GSHEETS else "excel"
    try:
        if USE_GSHEETS:
            sheets = _gs.get_order_sheet_names(GOOGLE_SHEETS_URL, _CREDS_PATH)
        else:
            if not os.path.exists(EXCEL_FILE):
                return jsonify({"sheets": [], "source": source, "error": "Файл не найден"})
            wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
            sheets = detect_order_sheets(wb)
            wb.close()
        return jsonify({"sheets": sheets, "source": source})
    except Exception as e:
        return jsonify({"sheets": [], "source": source, "error": str(e)}), 500


@app.route("/api/sheets")
def api_sheets():
    """Оставлен для совместимости."""
    return api_meta()


@app.route("/api/refresh", methods=["POST"])
def api_refresh():
    """Сбрасывает кеш и возвращает свежие данные из источника."""
    try:
        if USE_GSHEETS:
            _gs.invalidate_cache(GOOGLE_SHEETS_URL)
            source = "Google Sheets"
        else:
            source = "Excel"
        return jsonify({"ok": True, "message": f"Кеш сброшен, данные обновятся из {source}"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/salary")
def api_salary():
    """
    GET /api/salary?from=2026-04-01&to=2026-04-28
    Считает ЗП по всем листам за указанный период.
    """
    try:
        today = date.today()
        from_str = request.args.get("from")
        to_str   = request.args.get("to")

        if from_str:
            date_from = datetime.strptime(from_str, "%Y-%m-%d").date()
        else:
            date_from = date(today.year, today.month, 1)

        if to_str:
            date_to = datetime.strptime(to_str, "%Y-%m-%d").date()
        else:
            date_to = today

        if date_from > date_to:
            return jsonify({"error": "Дата «с» не может быть позже даты «по»"}), 400

        data = calculate_salary(date_from, date_to)
        return jsonify(data)

    except ValueError as e:
        return jsonify({"error": f"Неверный формат даты: {e}"}), 400
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/transfer", methods=["POST"])
def api_transfer():
    """
    Переносит незавершённые заказы из одного листа в другой.
    Body JSON: {"from_sheet": "Апрель_2026", "to_sheet": "Май_2026"}

    Работает в обоих режимах: Excel и Google Sheets.
    """
    try:
        body = request.get_json(force=True) or {}
        from_sheet = (body.get("from_sheet") or "").strip()
        to_sheet   = (body.get("to_sheet") or "").strip()

        if not from_sheet or not to_sheet:
            return jsonify({"error": "Укажи from_sheet и to_sheet"}), 400
        if from_sheet == to_sheet:
            return jsonify({"error": "Исходный и целевой листы совпадают"}), 400

        # ── Google Sheets режим ───────────────────────────────────────────────
        if USE_GSHEETS:
            count = _gs.transfer_orders(GOOGLE_SHEETS_URL, _CREDS_PATH, from_sheet, to_sheet)
            if count == 0:
                return jsonify({"transferred": 0, "message": "Незавершённых заказов нет"})
            return jsonify({
                "transferred": count,
                "from_sheet":  from_sheet,
                "to_sheet":    to_sheet,
                "message":     f"Перенесено {count} заказов из «{from_sheet}» в «{to_sheet}» (Google Sheets)",
            })

        # ── Локальный Excel режим ─────────────────────────────────────────────
        wb = openpyxl.load_workbook(EXCEL_FILE)

        if from_sheet not in wb.sheetnames:
            return jsonify({"error": f"Лист «{from_sheet}» не найден"}), 404

        if to_sheet not in wb.sheetnames:
            create_month_sheet(wb, to_sheet)

        ws_from = wb[from_sheet]
        ws_to   = wb[to_sheet]

        rows_to_move      = []
        row_nums_to_delete = []

        for row_num, row in enumerate(ws_from.iter_rows(min_row=2, values_only=False), start=2):
            vals = [c.value for c in row]
            if vals[0] is None and (len(vals) < 3 or vals[2] is None):
                continue
            if isinstance(vals[0], str) and vals[0].startswith("←"):
                continue
            if len(vals) < 18 or vals[17] is None:
                rows_to_move.append(vals)
                row_nums_to_delete.append(row_num)

        if not rows_to_move:
            return jsonify({"transferred": 0, "message": "Незавершённых заказов нет"})

        next_row = 2 if (ws_to.max_row <= 1 or ws_to.cell(2, 1).value is None) else ws_to.max_row + 1
        DATE_COLS_IDX = {0, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 17}

        for row_data in rows_to_move:
            for ci, val in enumerate(row_data[:19], start=1):
                cell = ws_to.cell(next_row, ci, value=val)
                cell.border = _border()
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(
                    horizontal="left" if ci in (3, 4, 5, 19) else "center",
                    vertical="center",
                )
                if (ci - 1) in DATE_COLS_IDX and isinstance(val, datetime):
                    cell.number_format = "DD.MM.YY"
            next_row += 1

        for row_num in reversed(row_nums_to_delete):
            ws_from.delete_rows(row_num)

        wb.save(EXCEL_FILE)

        return jsonify({
            "transferred": len(rows_to_move),
            "from_sheet":  from_sheet,
            "to_sheet":    to_sheet,
            "message":     f"Перенесено {len(rows_to_move)} заказов из «{from_sheet}» в «{to_sheet}»",
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


if __name__ == "__main__":
    print("=" * 50)
    print("  SignaturePro — Dashboard ZP")
    print("=" * 50)
    if not os.path.exists(EXCEL_FILE):
        print(f"\n  Файл не найден: {EXCEL_FILE}")
        print("  Сначала запусти: python create_orders_table.py\n")
    else:
        print(f"\n  Файл данных: {EXCEL_FILE}")
    print("\n  Открой в браузере: http://localhost:5000\n")
    app.run(debug=False, host="0.0.0.0", port=5000)
